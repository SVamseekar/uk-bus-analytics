#!/usr/bin/env python3
"""
Download UK Business Counts from ONS and convert to CSV
"""
import requests
import openpyxl
import csv
from pathlib import Path
from loguru import logger

# ONS URL for business counts XLSX
BUSINESS_COUNTS_URL = 'https://www.ons.gov.uk/file?uri=/businessindustryandtrade/business/activitysizeandlocation/datasets/ukbusinesscountslocalunitsby4digitsicindustryandemploymentsizeband/current/ukbc24lsoa.xlsx'

DEMOGRAPHICS_DIR = Path("data/raw/demographics")
OUTPUT_CSV = DEMOGRAPHICS_DIR / "business_counts.csv"
TEMP_XLSX = DEMOGRAPHICS_DIR / "business_counts_temp.xlsx"

def main():
    logger.info("Downloading UK Business Counts from ONS...")
    logger.info(f"URL: {BUSINESS_COUNTS_URL}")

    # Download XLSX file
    logger.info("Downloading XLSX file...")
    response = requests.get(BUSINESS_COUNTS_URL, timeout=300, stream=True)

    if response.status_code != 200:
        logger.error(f"Failed to download: HTTP {response.status_code}")
        return

    # Save XLSX
    with open(TEMP_XLSX, 'wb') as f:
        for chunk in response.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)

    file_size = TEMP_XLSX.stat().st_size
    logger.success(f"Downloaded: {file_size:,} bytes")

    # Convert to CSV
    logger.info("Converting XLSX to CSV...")
    wb = openpyxl.load_workbook(TEMP_XLSX)

    # Try to find the sheet with LSOA data
    logger.info(f"Available sheets: {wb.sheetnames}")

    # Use the first sheet or find one with 'LSOA' in name
    ws = None
    for sheet_name in wb.sheetnames:
        if 'LSOA' in sheet_name.upper() or 'MSOA' in sheet_name.upper():
            ws = wb[sheet_name]
            logger.info(f"Using sheet: {sheet_name}")
            break

    if ws is None:
        ws = wb.active
        logger.info(f"Using active sheet: {ws.title}")

    # Write to CSV
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                writer.writerow(row)
                row_count += 1

    # Clean up temp file
    TEMP_XLSX.unlink()

    output_size = OUTPUT_CSV.stat().st_size
    logger.success(f"✓ Converted to CSV: {OUTPUT_CSV}")
    logger.success(f"✓ {row_count:,} rows, {output_size:,} bytes")

    # Show first few rows
    import pandas as pd
    df = pd.read_csv(OUTPUT_CSV, nrows=5)
    logger.info(f"Columns: {list(df.columns)}")
    logger.info(f"Preview:\n{df.head()}")

if __name__ == "__main__":
    main()
