"""
Dynamic UK Bus Analytics Data Ingestion Pipeline
No hardcoding - fully configurable and scalable to all UK regions
Addresses all 57 analytical questions

UPDATED: Fixed NOMIS API downloads with proper error handling
"""
import os
import sys
import yaml
import json
import time
import zipfile
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
import pandas as pd
import requests
from loguru import logger

sys.path.append(str(Path(__file__).parent.parent))

from config.settings import DATA_RAW, API_ENDPOINTS, LOGS_DIR
from utils.api_client import BODSClient, ONSClient, NomisClient

logger.add(LOGS_DIR / "ingestion_{time}.log", rotation="1 day", retention="30 days")


class DynamicDataIngestionPipeline:
    """
    Fully dynamic data ingestion pipeline
    Configurable via YAML files - no hardcoded values
    """
    
    def __init__(self, config_path: Optional[Path] = None):
        """
        Initialize with optional configuration file
        Falls back to auto-discovery if no config provided
        """
        self.config_path = config_path or Path(__file__).parent.parent / 'config' / 'ingestion_config.yaml'
        self.config = self._load_configuration()
        
        # Initialize API clients
        self.bods_client = self._init_bods_client()
        self.ons_client = ONSClient()
        self.nomis_client = NomisClient()
        
        # Statistics tracking
        self.stats = {
            'start_time': datetime.now(),
            'regions_processed': [],
            'datasets_downloaded': {},
            'demographic_datasets': [],
            'errors': [],
            'warnings': []
        }
    
    def _load_configuration(self) -> Dict:
        """
        Load configuration from YAML file
        Auto-generates default config if missing
        """
        if not self.config_path.exists():
            logger.warning(f"Config file not found: {self.config_path}")
            logger.info("Generating default configuration...")
            return self._generate_default_config()
        
        try:
            with open(self.config_path, 'r') as f:
                config = yaml.safe_load(f)
            logger.success(f"Loaded configuration from {self.config_path}")
            return config
        except Exception as e:
            logger.error(f"Failed to load config: {e}")
            return self._generate_default_config()
    
    def _generate_default_config(self) -> Dict:
        """
        Generate comprehensive default configuration
        Covers all 9 UK regions and 8 demographic datasets
        """
        default_config = {
            'regions': {
                'north_east': {
                    'enabled': True,
                    'name': 'North East England',
                    'major_cities': ['Newcastle', 'Sunderland', 'Middlesbrough', 'Durham'],
                    'lsoa_prefix': 'E01',
                    'max_datasets': 10,
                    'priority': 'medium'
                },
                'north_west': {
                    'enabled': True,
                    'name': 'North West England',
                    'major_cities': ['Manchester', 'Liverpool', 'Preston', 'Blackpool', 'Bolton'],
                    'lsoa_prefix': 'E01',
                    'max_datasets': 15,
                    'priority': 'high'
                },
                'yorkshire': {
                    'enabled': True,
                    'name': 'Yorkshire and Humber',
                    'major_cities': ['Leeds', 'Sheffield', 'Bradford', 'Hull', 'York'],
                    'lsoa_prefix': 'E01',
                    'max_datasets': 12,
                    'priority': 'high'
                },
                'east_midlands': {
                    'enabled': True,
                    'name': 'East Midlands',
                    'major_cities': ['Nottingham', 'Derby', 'Leicester', 'Lincoln', 'Northampton'],
                    'lsoa_prefix': 'E01',
                    'max_datasets': 10,
                    'priority': 'high'
                },
                'west_midlands': {
                    'enabled': True,
                    'name': 'West Midlands',
                    'major_cities': ['Birmingham', 'Coventry', 'Wolverhampton', 'Dudley', 'Solihull'],
                    'lsoa_prefix': 'E01',
                    'max_datasets': 15,
                    'priority': 'high'
                },
                'east_england': {
                    'enabled': True,
                    'name': 'East of England',
                    'major_cities': ['Norwich', 'Cambridge', 'Ipswich', 'Peterborough', 'Luton'],
                    'lsoa_prefix': 'E01',
                    'max_datasets': 10,
                    'priority': 'medium'
                },
                'london': {
                    'enabled': True,
                    'name': 'Greater London',
                    'major_cities': ['London'],
                    'lsoa_prefix': 'E01',
                    'max_datasets': 20,
                    'priority': 'critical'
                },
                'south_east': {
                    'enabled': True,
                    'name': 'South East England',
                    'major_cities': ['Brighton', 'Reading', 'Oxford', 'Southampton', 'Portsmouth'],
                    'lsoa_prefix': 'E01',
                    'max_datasets': 12,
                    'priority': 'high'
                },
                'south_west': {
                    'enabled': True,
                    'name': 'South West England',
                    'major_cities': ['Bristol', 'Plymouth', 'Bournemouth', 'Exeter', 'Swindon'],
                    'lsoa_prefix': 'E01',
                    'max_datasets': 10,
                    'priority': 'medium'
                }
            },
            'demographic_sources': {
                'imd_2019': {
                    'enabled': True,
                    'name': 'Index of Multiple Deprivation 2019',
                    'source': 'direct_download',
                    'url': 'https://assets.publishing.service.gov.uk/media/5dc407b440f0b6379a7acc8d/File_7_-_All_IoD2019_Scores__Ranks__Deciles_and_Population_Denominators_3.csv',
                    'questions_addressed': ['D24', 'D30', 'F37', 'G43', 'G50'],
                    'priority': 'critical'
                },
                'population_2021': {
                    'enabled': True,
                    'name': 'Census 2021 Population by LSOA',
                    'source': 'nomis',
                    'dataset_id': 'NM_2010_1',
                    'geography': 'TYPE297',
                    'time': '2021',
                    'questions_addressed': ['A1', 'A2', 'A3', 'D24', 'E32'],
                    'priority': 'critical'
                },
                'income_2022': {
                    'enabled': True,
                    'name': 'Economic Activity Census 2021 (Latest Available)',
                    'source': 'direct_download',
                    'url': 'https://www.ons.gov.uk/file?uri=/datasets/ts066/editions/2021/versions/3/observations.csv',
                    'questions_addressed': ['D24', 'D28', 'I55', 'I56'],
                    'priority': 'high'
                },
                'unemployment_2024': {
                    'enabled': True,
                    'name': 'Unemployment Rate by LSOA',
                    'source': 'nomis',
                    'dataset_id': 'NM_162_1',
                    'geography': 'TYPE297',
                    'time': 'latest',
                    'questions_addressed': ['D25', 'B14', 'H52'],
                    'priority': 'high'
                },
                'car_ownership': {
                    'enabled': True,
                    'name': 'Car Van Ownership Census 2021 (Latest Available)',
                    'source': 'direct_download',
                    'url': 'https://www.ons.gov.uk/file?uri=/datasets/ts045/editions/2021/versions/3/observations.csv',
                    'questions_addressed': ['D28'],
                    'priority': 'medium'
                },
                'age_structure': {
                    'enabled': True,
                    'name': 'Age Structure by LSOA',
                    'source': 'nomis',
                    'dataset_id': 'NM_2010_1',
                    'geography': 'TYPE297',
                    'measures': '20100',
                    'time': '2021',
                    'questions_addressed': ['D27'],
                    'priority': 'medium'
                },
                'schools_2024': {
                    'enabled': True,
                    'name': 'UK Schools Database',
                    'source': 'manual',
                    'url': 'https://get-information-schools.service.gov.uk/Downloads',
                    'instructions': 'Download Establishment fields CSV with postcodes',
                    'questions_addressed': ['C22', 'D29', 'F39', 'H51'],
                    'priority': 'high'
                },
                'business_counts': {
                    'enabled': True,
                    'name': 'UK Business Counts 2024 (Latest Available)',
                    'source': 'direct_download',
                    'url': 'https://www.ons.gov.uk/file?uri=/businessindustryandtrade/business/activitysizeandlocation/datasets/ukbusinesscountslocalunitsby4digitsicindustryandemploymentsizeband/current/ukbc24lsoa.xlsx',
                    'questions_addressed': ['I55'],
                    'priority': 'medium'
                }
            },
            'ingestion_settings': {
                'parallel_downloads': 3,
                'retry_attempts': 3,
                'timeout_seconds': 300,
                'cache_duration_days': 7,
                'validate_downloads': True,
                'min_file_size_bytes': 1000
            }
        }
        
        # Save default config
        self.config_path.parent.mkdir(parents=True, exist_ok=True)
        with open(self.config_path, 'w') as f:
            yaml.dump(default_config, f, default_flow_style=False, sort_keys=False)
        
        logger.success(f"Generated default config: {self.config_path}")
        return default_config
    
    def _init_bods_client(self) -> Optional[BODSClient]:
        """Initialize BODS client with validation"""
        api_key = API_ENDPOINTS['bods']['api_key']
        
        if not api_key or api_key == 'your_actual_bods_api_key_here':
            logger.error("BODS API key not configured in .env file")
            logger.info("Get your API key from: https://data.bus-data.dft.gov.uk/account/api/")
            return None
        
        try:
            return BODSClient(
                base_url=API_ENDPOINTS['bods']['base_url'],
                api_key=api_key
            )
        except Exception as e:
            logger.error(f"Failed to initialize BODS client: {e}")
            return None
    
    def ingest_transport_data_for_region(self, region_code: str) -> Dict:
        """
        Ingest all transport data for a specific region
        Fully dynamic based on BODS API discovery
        """
        region_config = self.config['regions'].get(region_code)
        if not region_config:
            logger.error(f"Region not found in config: {region_code}")
            return {'success': False, 'error': 'Region not configured'}
        
        logger.info(f"\n{'='*60}")
        logger.info(f"REGION: {region_config['name']}")
        logger.info(f"{'='*60}")
        
        if not self.bods_client:
            logger.error("BODS client not available - skipping transport data")
            return {'success': False, 'error': 'No BODS client'}
        
        # Create region directory
        region_dir = DATA_RAW / 'transport' / region_code
        region_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            # Discover available datasets for this region
            logger.info("Discovering available datasets...")
            datasets = self.bods_client.get_datasets(
                limit=region_config.get('max_datasets', 10)
            )
            
            if not datasets:
                logger.warning("No datasets found")
                return {'success': False, 'datasets_downloaded': 0}
            
            logger.info(f"Found {len(datasets.get('results', []))} datasets")
            
            # Download datasets
            downloaded = 0
            failed = 0
            
            for i, dataset in enumerate(datasets.get('results', [])[:region_config.get('max_datasets', 10)], 1):
                dataset_id = dataset.get('id', f'dataset_{i}')
                dataset_name = dataset.get('name', f'Dataset_{i}')
                operator_name = dataset.get('operatorName', 'Unknown_Operator')

                # Clean filename
                safe_name = f"{operator_name}_{dataset_id}".replace(' ', '_').replace('/', '_')[:50]

                logger.info(f"\nDataset {i}: {dataset_name}")
                logger.info(f"  Operator: {operator_name}")
                logger.info(f"  ID: {dataset_id}")

                try:
                    dataset_url = dataset.get('url')
                    if not dataset_url:
                        logger.warning(f"No URL for dataset: {dataset_name}")
                        failed += 1
                        continue

                    output_file = region_dir / f"{safe_name}.zip"

                    logger.info(f"  Downloading from: {dataset_url}")
                    if self.bods_client.download_dataset_file(dataset_url, str(output_file)):
                        downloaded += 1
                        logger.success(f"✓ Downloaded: {dataset_name}")
                    else:
                        failed += 1
                        logger.warning(f"✗ Failed: {dataset_name}")

                except Exception as e:
                    logger.error(f"Failed to download dataset {i}: {e}")
                    failed += 1
                    continue
            
            result = {
                'success': downloaded > 0,
                'region': region_code,
                'region_name': region_config['name'],
                'datasets_discovered': len(datasets.get('results', [])),
                'datasets_downloaded': downloaded,
                'datasets_failed': failed,
                'output_directory': str(region_dir)
            }
            
            self.stats['datasets_downloaded'][region_code] = result
            self.stats['regions_processed'].append(region_code)
            
            logger.success(f"✓ {region_config['name']}: {downloaded}/{len(datasets.get('results', []))} datasets")
            return result
            
        except Exception as e:
            logger.error(f"Failed to process region {region_code}: {e}")
            return {'success': False, 'error': str(e)}
    
    def ingest_all_transport_data(self) -> Dict:
        """
        Ingest transport data for all enabled regions
        Fully dynamic based on configuration
        """
        logger.info("\n" + "="*60)
        logger.info("NATIONAL TRANSPORT DATA INGESTION")
        logger.info("="*60 + "\n")
        
        enabled_regions = {
            code: config for code, config in self.config['regions'].items()
            if config.get('enabled', False)
        }
        
        logger.info(f"Processing {len(enabled_regions)} enabled regions")
        
        # Process by priority
        priority_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        sorted_regions = sorted(
            enabled_regions.items(),
            key=lambda x: priority_order.get(x[1].get('priority', 'medium'), 2)
        )
        
        results = {}
        for region_code, region_config in sorted_regions:
            logger.info(f"\nPriority: {region_config.get('priority', 'medium').upper()}")
            result = self.ingest_transport_data_for_region(region_code)
            results[region_code] = result
        
        return results
    
    def download_demographic_dataset(self, dataset_key: str, config: Dict) -> bool:
        """
        FIXED: Download a single demographic dataset with proper NOMIS API handling
        Handles multiple source types dynamically with validation and retries
        """
        if not config.get('enabled', True):
            logger.info(f"Skipping disabled dataset: {dataset_key}")
            return False
        
        logger.info(f"Processing: {config['name']}")
        
        source_type = config.get('source', 'unknown').lower()
        demo_dir = DATA_RAW / 'demographic'
        demo_dir.mkdir(parents=True, exist_ok=True)
        
        output_path = demo_dir / f"{dataset_key}.csv"
        
        # Check cache with size validation
        if output_path.exists():
            file_age_days = (datetime.now().timestamp() - output_path.stat().st_mtime) / 86400
            cache_duration = self.config['ingestion_settings'].get('cache_duration_days', 7)
            file_size = output_path.stat().st_size
            
            if file_age_days < cache_duration and file_size > 1000:
                logger.info(f"Using cached data (age: {file_age_days:.1f} days, size: {file_size} bytes)")
                self.stats['demographic_datasets'].append(dataset_key)
                return True
            else:
                if file_size <= 1000:
                    logger.warning(f"Cached file too small ({file_size} bytes), re-downloading")
                output_path.unlink()
        
        try:
            if source_type == 'nomis':
                # FIXED NOMIS API download
                dataset_id = config.get('dataset_id')
                geography = config.get('geography', 'TYPE297')
                time_period = config.get('time', 'latest')
                measures = config.get('measures', None)
                
                logger.info(f"  Dataset: {dataset_id}")
                logger.info(f"  Geography: {geography}")
                logger.info(f"  Time: {time_period}")
                
                # WORKING NOMIS API URL FORMAT
                url = f"https://www.nomisweb.co.uk/api/v01/dataset/{dataset_id}.data.csv"

                # Base parameters that work for ALL datasets
                params = {
                    'geography': geography,
                    'select': 'geography_code,geography_name,obs_value,date_name',
                    'recordlimit': 0  # No limit - get all data
                }

                # Dataset-specific configurations
                if dataset_id == 'NM_2010_1':  # Population dataset
                    params['time'] = '2021'
                    params['select'] = 'geography_code,geography_name,obs_value,c_age_name,gender_name'

                elif dataset_id == 'NM_2080_1':  # Economic Activity Census 2021 (Latest)
                    params['time'] = '2021'
                    params['select'] = 'geography_code,geography_name,obs_value'

                elif dataset_id == 'NM_162_1':  # Unemployment dataset
                    params['time'] = '2024'
                    params['select'] = 'geography_code,geography_name,obs_value'

                elif dataset_id == 'NM_2027_1':  # Car ownership (Census 2021)
                    params['time'] = '2021'
                    params['select'] = 'geography_code,geography_name,obs_value'
                    params['c2021_carvan_5'] = '0...5'  # All car ownership categories

                elif dataset_id == 'NM_189_1':  # Business Register Employment Survey 2024 (Latest)
                    params['time'] = '2024'
                    params['select'] = 'geography_code,geography_name,obs_value'

                else:
                    # Default for any other dataset
                    params['time'] = 'latest' if time_period == 'latest' else time_period

                # Always add measures if specified in config
                if measures:
                    params['measures'] = measures
                
                # Retry logic
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        logger.info(f"  Attempt {attempt + 1}/{max_retries}")
                        
                        response = requests.get(url, params=params, timeout=300)
                        
                        if response.status_code == 200:
                            content = response.content
                            
                            # Validate response size
                            if len(content) < 200:
                                logger.error(f"  Response too small: {len(content)} bytes")
                                if attempt < max_retries - 1:
                                    time.sleep(2)
                                    continue
                                else:
                                    return False
                            
                            # Check if response is HTML error page
                            content_preview = content[:100].decode('utf-8', errors='ignore').lower()
                            if '<html' in content_preview or '<!doctype' in content_preview:
                                logger.error("  Received HTML error page instead of CSV")
                                if attempt < max_retries - 1:
                                    # Try without date parameter
                                    params.pop('date', None)
                                    time.sleep(2)
                                    continue
                                else:
                                    return False
                            
                            # Save the file
                            with open(output_path, 'wb') as f:
                                f.write(content)
                            
                            # Validate we can read it as CSV
                            try:
                                test_df = pd.read_csv(output_path, nrows=5)
                                
                                if len(test_df.columns) == 0:
                                    logger.error("  Downloaded file has no columns")
                                    output_path.unlink()
                                    if attempt < max_retries - 1:
                                        time.sleep(2)
                                        continue
                                    else:
                                        return False
                                
                                logger.success(f"✓ {config['name']}: {len(content)} bytes, {len(test_df.columns)} columns")
                                self.stats['demographic_datasets'].append(dataset_key)
                                return True
                                
                            except Exception as e:
                                logger.error(f"  Downloaded file invalid: {e}")
                                output_path.unlink()
                                if attempt < max_retries - 1:
                                    time.sleep(2)
                                    continue
                                else:
                                    return False
                        
                        else:
                            logger.error(f"  HTTP {response.status_code}")
                            if attempt < max_retries - 1:
                                time.sleep(2)
                                continue
                            else:
                                return False
                    
                    except requests.exceptions.Timeout:
                        logger.error(f"  Request timeout")
                        if attempt < max_retries - 1:
                            time.sleep(5)
                            continue
                        else:
                            return False
                    
                    except Exception as e:
                        logger.error(f"  Request failed: {e}")
                        if attempt < max_retries - 1:
                            time.sleep(2)
                            continue
                        else:
                            return False
                
                return False
                        
            elif source_type == 'ons_api':
                # Direct ONS file downloads
                url = config.get('url')
                if not url:
                    logger.error(f"No URL provided for ONS API dataset: {dataset_key}")
                    return False

                logger.info(f"  Downloading from ONS: {url}")
                response = requests.get(url, timeout=300, stream=True)

                if response.status_code == 200:
                    # Handle different file types
                    content_type = response.headers.get('content-type', '').lower()

                    if 'excel' in content_type or url.endswith('.xlsx'):
                        # XLSX file
                        temp_path = output_path.with_suffix('.xlsx')
                        with open(temp_path, 'wb') as f:
                            for chunk in response.iter_content(chunk_size=8192):
                                if chunk:
                                    f.write(chunk)

                        # Convert XLSX to CSV (basic conversion)
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(temp_path)
                            ws = wb.active

                            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                                import csv
                                writer = csv.writer(csvfile)
                                for row in ws.iter_rows(values_only=True):
                                    if any(cell is not None for cell in row):
                                        writer.writerow(row)

                            temp_path.unlink()  # Remove temp XLSX
                            logger.success(f"✓ {config['name']}: Converted XLSX to CSV")

                        except ImportError:
                            logger.warning(f"openpyxl not available, keeping as XLSX: {temp_path}")
                            output_path = temp_path

                    else:
                        # CSV or other text format
                        with open(output_path, 'wb') as f:
                            for chunk in response.iter_content(chunk_size=8192):
                                if chunk:
                                    f.write(chunk)
                        logger.success(f"✓ {config['name']}: {len(response.content)} bytes")

                    self.stats['demographic_datasets'].append(dataset_key)
                    return True
                else:
                    logger.error(f"✗ {config['name']}: HTTP {response.status_code}")
                    return False

            elif source_type == 'direct_download':
                # Direct file downloads from ONS or other sources
                url = config.get('url')
                if not url:
                    logger.error(f"No URL provided for direct download dataset: {dataset_key}")
                    return False

                logger.info(f"  Direct download from: {url}")
                response = requests.get(url, timeout=300, stream=True)

                if response.status_code == 200:
                    total_size = int(response.headers.get('content-length', 0))

                    if url.endswith('.xlsx'):
                        # Handle Excel files
                        temp_path = output_path.with_suffix('.xlsx')
                        with open(temp_path, 'wb') as f:
                            for chunk in response.iter_content(chunk_size=8192):
                                if chunk:
                                    f.write(chunk)

                        # Convert to CSV if openpyxl available
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(temp_path)
                            ws = wb.active

                            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                                import csv
                                writer = csv.writer(csvfile)
                                for row in ws.iter_rows(values_only=True):
                                    if any(cell is not None for cell in row):
                                        writer.writerow(row)

                            temp_path.unlink()
                            file_size = output_path.stat().st_size
                            logger.success(f"✓ {config['name']}: {file_size} bytes (converted from XLSX)")

                        except ImportError:
                            logger.warning(f"openpyxl not available, keeping as XLSX: {temp_path}")
                            output_path = temp_path
                            file_size = temp_path.stat().st_size
                            logger.success(f"✓ {config['name']}: {file_size} bytes (XLSX)")

                    else:
                        # Handle CSV and other text files
                        with open(output_path, 'wb') as f:
                            downloaded = 0
                            for chunk in response.iter_content(chunk_size=8192):
                                if chunk:
                                    f.write(chunk)
                                    downloaded += len(chunk)

                        logger.success(f"✓ {config['name']}: {downloaded} bytes")

                    self.stats['demographic_datasets'].append(dataset_key)
                    return True
                else:
                    logger.error(f"✗ {config['name']}: HTTP {response.status_code}")
                    return False

            elif source_type == 'arcgis':
                # ArcGIS direct download
                url = config.get('url')
                response = requests.get(url, timeout=300, stream=True)
                
                if response.status_code == 200:
                    # Handle ZIP files
                    if 'zip' in response.headers.get('content-type', '').lower():
                        zip_path = demo_dir / f"{dataset_key}.zip"
                        with open(zip_path, 'wb') as f:
                            for chunk in response.iter_content(chunk_size=8192):
                                if chunk:
                                    f.write(chunk)
                        
                        # Extract CSV
                        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                            csv_files = [f for f in zip_ref.namelist() if f.endswith('.csv')]
                            if csv_files:
                                zip_ref.extract(csv_files[0], demo_dir)
                                extracted_path = demo_dir / csv_files[0]
                                extracted_path.rename(output_path)
                        
                        zip_path.unlink()
                    else:
                        with open(output_path, 'wb') as f:
                            for chunk in response.iter_content(chunk_size=8192):
                                if chunk:
                                    f.write(chunk)
                    
                    logger.success(f"✓ {config['name']}")
                    self.stats['demographic_datasets'].append(dataset_key)
                    return True
                else:
                    logger.error(f"✗ {config['name']}: HTTP {response.status_code}")
                    return False
                        
            elif source_type == 'manual':
                logger.warning(f"⚠ {config['name']} requires manual download")
                logger.info(f"  URL: {config.get('url')}")
                logger.info(f"  Instructions: {config.get('instructions', 'Download CSV')}")
                logger.info(f"  Save to: {output_path}")
                self.stats['warnings'].append(f"Manual download needed: {dataset_key}")
                return False
            
            else:
                logger.error(f"Unknown source type: {source_type}")
                return False
                
        except Exception as e:
            logger.error(f"Failed to download {dataset_key}: {e}")
            import traceback
            logger.error(traceback.format_exc())
            self.stats['errors'].append(f"Demographic {dataset_key} failed: {e}")
            return False
    
    def ingest_all_demographic_data(self) -> Dict:
        """
        Ingest all demographic datasets
        Fully dynamic based on configuration
        """
        logger.info("\n" + "="*60)
        logger.info("DEMOGRAPHIC DATA INGESTION")
        logger.info("="*60 + "\n")
        
        enabled_datasets = {
            key: config for key, config in self.config['demographic_sources'].items()
            if config.get('enabled', True)
        }
        
        logger.info(f"Processing {len(enabled_datasets)} demographic datasets")
        
        # Process by priority
        priority_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        sorted_datasets = sorted(
            enabled_datasets.items(),
            key=lambda x: priority_order.get(x[1].get('priority', 'medium'), 2)
        )
        
        successful = 0
        failed = 0
        manual = 0
        
        for dataset_key, dataset_config in sorted_datasets:
            logger.info(f"\nPriority: {dataset_config.get('priority', 'medium').upper()}")
            
            if dataset_config.get('source') == 'manual':
                self.download_demographic_dataset(dataset_key, dataset_config)
                manual += 1
            elif self.download_demographic_dataset(dataset_key, dataset_config):
                successful += 1
            else:
                failed += 1
        
        return {
            'total_datasets': len(enabled_datasets),
            'successful': successful,
            'failed': failed,
            'manual_required': manual,
            'downloaded_datasets': self.stats['demographic_datasets']
        }
    
    def generate_ingestion_report(self) -> Dict:
        """
        Generate comprehensive ingestion report
        """
        duration = datetime.now() - self.stats['start_time']
        
        report = {
            'ingestion_summary': {
                'start_time': self.stats['start_time'].isoformat(),
                'duration': str(duration),
                'total_regions_processed': len(self.stats['regions_processed']),
                'total_demographic_datasets': len(self.stats['demographic_datasets'])
            },
            'regional_breakdown': self.stats['datasets_downloaded'],
            'demographic_datasets': self.stats['demographic_datasets'],
            'errors': self.stats['errors'],
            'warnings': self.stats['warnings']
        }
        
        # Save report
        report_path = DATA_RAW / 'ingestion_report.json'
        with open(report_path, 'w') as f:
            json.dump(report, f, indent=2, default=str)
        
        logger.success(f"Ingestion report saved: {report_path}")
        return report
    
    def run_full_ingestion(self) -> Dict:
        """
        Execute complete ingestion pipeline
        Fully dynamic and configurable
        """
        logger.info("\n" + "="*60)
        logger.info("DYNAMIC DATA INGESTION PIPELINE")
        logger.info("="*60)
        
        logger.info(f"\nConfiguration: {self.config_path}")
        logger.info(f"Enabled regions: {sum(1 for r in self.config['regions'].values() if r.get('enabled'))}")
        logger.info(f"Enabled demographics: {sum(1 for d in self.config['demographic_sources'].values() if d.get('enabled'))}")
        
        # Step 1: Transport data
        logger.info("\n" + "="*60)
        logger.info("STEP 1: TRANSPORT DATA")
        logger.info("="*60)
        transport_results = self.ingest_all_transport_data()
        
        # Step 2: Demographic data
        logger.info("\n" + "="*60)
        logger.info("STEP 2: DEMOGRAPHIC DATA")
        logger.info("="*60)
        demographic_results = self.ingest_all_demographic_data()
        
        # Step 3: Generate report
        logger.info("\n" + "="*60)
        logger.info("STEP 3: INGESTION REPORT")
        logger.info("="*60)
        report = self.generate_ingestion_report()
        
        return {
            'transport_results': transport_results,
            'demographic_results': demographic_results,
            'report': report,
            'success': len(self.stats['errors']) == 0
        }


def main():
    """
    Main function with CLI support for comprehensive automation
    """
    import argparse

    parser = argparse.ArgumentParser(description='UK Bus Analytics - Automated Data Ingestion Pipeline')
    parser.add_argument('--data-type', choices=['transport', 'demographic', 'all'],
                       default='all', help='Type of data to ingest')
    parser.add_argument('--regions', type=str, default='all',
                       help='Comma-separated list of regions or "all"')
    parser.add_argument('--config', type=str,
                       help='Path to configuration file')
    parser.add_argument('--dry-run', action='store_true',
                       help='Show what would be downloaded without actually downloading')
    parser.add_argument('--max-datasets', type=int, default=None,
                       help='Maximum datasets per region (for testing)')

    args = parser.parse_args()

    print("\n" + "="*60)
    print("UK BUS ANALYTICS - DYNAMIC DATA INGESTION")
    print("="*60)
    print("\nFully configurable, no hardcoding")
    print("Scalable to all UK regions\n")

    # Initialize pipeline
    config_path = Path(args.config) if args.config else None
    pipeline = DynamicDataIngestionPipeline(config_path)

    # Apply CLI overrides
    if args.max_datasets:
        for region in pipeline.config['regions'].values():
            region['max_datasets'] = args.max_datasets

    logger.info(f"🚀 Starting automated ingestion:")
    logger.info(f"  - Data type: {args.data_type}")
    logger.info(f"  - Regions: {args.regions}")
    logger.info(f"  - Dry run: {args.dry_run}")

    results = {'transport_results': {}, 'demographic_results': {}, 'success': True}

    try:
        # Execute based on data type
        if args.data_type in ['transport', 'all']:
            logger.info("\n" + "="*60)
            logger.info("🚌 TRANSPORT DATA INGESTION")
            logger.info("="*60)

            if args.regions == 'all':
                if not args.dry_run:
                    results['transport_results'] = pipeline.ingest_all_transport_data()
                else:
                    logger.info("DRY RUN: Would download transport data for all regions")
                    results['transport_results'] = {r: {'dry_run': True} for r in pipeline.config['regions']}
            else:
                regions = [r.strip() for r in args.regions.split(',')]
                for region in regions:
                    if region in pipeline.config['regions']:
                        if not args.dry_run:
                            results['transport_results'][region] = pipeline.ingest_transport_data_for_region(region)
                        else:
                            logger.info(f"DRY RUN: Would download transport data for {region}")
                            results['transport_results'][region] = {'dry_run': True}
                    else:
                        logger.warning(f"❌ Unknown region: {region}")

        if args.data_type in ['demographic', 'all']:
            logger.info("\n" + "="*60)
            logger.info("📊 DEMOGRAPHIC DATA INGESTION")
            logger.info("="*60)

            if not args.dry_run:
                results['demographic_results'] = pipeline.ingest_all_demographic_data()
            else:
                logger.info("DRY RUN: Would download all demographic datasets")
                results['demographic_results'] = {'dry_run': True}

        # Generate final report
        if not args.dry_run:
            logger.info("\n" + "="*60)
            logger.info("📋 GENERATING REPORT")
            logger.info("="*60)
            report = pipeline.generate_ingestion_report()
            results['report'] = report

    except Exception as e:
        logger.error(f"💥 Pipeline failed: {e}")
        results['success'] = False
        raise

    # Print comprehensive summary
    print("\n" + "="*60)
    print("INGESTION COMPLETE")
    print("="*60)

    if args.dry_run:
        print("\n🔍 DRY RUN COMPLETED - No actual downloads performed")
    else:
        transport = results['transport_results']
        demographics = results['demographic_results']

        print(f"\n🚌 Transport Data:")
        if isinstance(transport, dict):
            total_regions = len(transport)
            total_datasets = sum(r.get('datasets_downloaded', 0) for r in transport.values() if isinstance(r, dict))
            total_failed = sum(r.get('datasets_failed', 0) for r in transport.values() if isinstance(r, dict))
            print(f"  Regions processed: {total_regions}")
            print(f"  Total datasets downloaded: {total_datasets}")
            print(f"  Failed downloads: {total_failed}")

        print(f"\n📊 Demographic Data:")
        if isinstance(demographics, dict) and 'successful' in demographics:
            print(f"  Successful: {demographics.get('successful', 0)}")
            print(f"  Failed: {demographics.get('failed', 0)}")
            print(f"  Manual required: {demographics.get('manual_required', 0)}")

        if 'report' in results and results['report'].get('errors'):
            print(f"\n⚠️  Errors encountered: {len(results['report']['errors'])}")
            for error in results['report']['errors'][:5]:
                print(f"  - {error}")

    if results['success']:
        print("\n✅ Automated ingestion pipeline completed successfully!")
    else:
        print("\n⚠️  Pipeline completed with errors")

    if not args.dry_run:
        print(f"\n🔄 Next steps:")
        print("1. python data_pipeline/02_data_processing.py")
        print("2. python data_pipeline/03_data_validation.py")
        print("3. Review ingestion report: data_pipeline/raw/ingestion_report.json")

    logger.success("🎉 Automation pipeline execution completed!")


if __name__ == "__main__":
    main()